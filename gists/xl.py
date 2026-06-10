"""Excel — openpyxl workbooks: previews, islands, table-aware writes (llmxlsx core)."""

import sys

__repld_deps__ = ["openpyxl>=3.1.5"]
__repld_usage__ = "xl = XL('report.xlsx'); print(xl.preview('Sheet1'))"

# llmxlsx core (query/sampling/formulas/parser) needs only openpyxl + stdlib.
# Import from source so edits there land on next gist reload — no install.
_LLMXLSX_SRC = "/home/fredrik/Projects/private/llmxlsx/src"
if _LLMXLSX_SRC not in sys.path:
    sys.path.insert(0, _LLMXLSX_SRC)

from llmxlsx.query import ExcelQuery  # noqa: E402


class XL(ExcelQuery):
    """ExcelQuery with cross-pane conveniences.

    XL(path)            open existing workbook
    XL.create(path)     new workbook (refuses to overwrite)
    .preview(sheet)     Excel-style grid as string — print() it
    .island('S!A1:D9')  column samples + patterns for a range
    Inherited from ExcelQuery (not in introspection, full docs in llmxlsx):
    .append_row('S!A8', {'A': ...})  row templating — column-LETTER keys,
                                     unknown keys silently ignored
    .set / .write_row / .write_col / .cell / .clear_range / .save
    .create_sheet / .delete_sheet / .rename_sheet / .move_sheet
    """

    def __init__(self, file_path: str) -> None:
        """Open an existing workbook. Use XL.create(path) for a new one."""
        super().__init__(file_path)

    def preview(
        self,
        sheet_name: str,
        row_limit: int | None = None,
        col_limit: int | None = None,
    ) -> str:
        """Excel-style grid of a sheet as plain text -> print() it."""
        return super().preview(sheet_name, row_limit, col_limit)

    def island(self, range_spec: str) -> dict:
        """Column samples + patterns for a range -> {range, size, cols: {A: {samples, pattern}}}."""
        return super().island(range_spec)

    def autofit(self, sheet: str, min_width: int = 8, max_width: int = 60) -> None:
        """Set column widths from content length (openpyxl has no native autofit)."""
        from openpyxl.utils import get_column_letter

        ws = self._get_worksheet(sheet)
        for col in range(1, (ws.max_column or 1) + 1):
            longest = max(
                (
                    len(str(c.value))
                    for c in ws[get_column_letter(col)]
                    if c.value is not None
                ),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col)].width = max(
                min_width, min(longest + 2, max_width)
            )

    def append(self, table: str, record: dict, save: bool = True) -> str:
        """Header-keyed append to a named Excel table -> new row range 'A22:N22'.

        Copies formulas/formatting from the last data row, then expands the
        table ref, auto-filter, data validations, and conditional formatting.
        """
        from llmxlsx.xlkit.managers.table_manager import TableManager

        tm = TableManager(self)
        t, ws = tm.get_table_by_name(table)
        rng = tm.append_to_table(t, ws, record)
        if save:
            self.save()
        return rng

    def dump(
        self,
        records: list[dict],
        sheet: str = "Sheet1",
        table: str | None = None,
        save: bool = True,
    ) -> str:
        """Write list-of-dicts as header + rows at A1 -> written range 'Sheet1!A1:C11'.

        Headers come from the union of record keys, ordered by first appearance.
        With table=, registers an Excel table over the range (banded style,
        enables .append()'s formula-expanding behavior). Autofits columns.
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        headers: list[str] = []
        for r in records:
            for k in r:
                if k not in headers:
                    headers.append(k)
        if not headers:
            raise ValueError("records is empty or has no keys")

        if sheet not in self.workbook.sheetnames:
            self.create_sheet(sheet)
        ws = self._get_worksheet(sheet)

        ws.append(headers)
        for r in records:
            ws.append([r.get(h) for h in headers])

        ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
        if table:
            t = Table(displayName=table, ref=ref)
            t.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True
            )
            ws.add_table(t)
        self.autofit(sheet)
        if save:
            self.save()
        return f"{sheet}!{ref}"
